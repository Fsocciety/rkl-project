import openpyxl


def handle_date(date):
    temp = date.split('.')
    return temp[2] + '-' + temp[1] + '-' + temp[0]


def read(file_location):
    wb = openpyxl.load_workbook(file_location)
    ws = wb.active
    reports = []
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value is not None:
            report = (
                int(ws.cell(row=row, column=2).value),
                handle_date(ws.cell(row=row, column=3).value),
                ws.cell(row=row, column=4).value,
                ws.cell(row=row, column=5).value,
                ws.cell(row=row, column=6).value,
                ws.cell(row=row, column=7).value,
                ws.cell(row=row, column=8).value,
                ws.cell(row=row, column=9).value,
                ws.cell(row=row, column=10).value,
                int(ws.cell(row=row, column=11).value),
                int(ws.cell(row=row, column=12).value),
                int(ws.cell(row=row, column=13).value),
            )
            reports.append(report)
    print(reports)
    return reports